from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

from agents.config import JOBS_PATH, TRACKER_PATH, ensure_data_dirs
from agents.evaluation import EvaluationAgent
from agents.models import EvaluationResult, JobListing, RoutingDecision, ApplicationStatus, save_jobs, load_jobs


def normalize_job_url(url: str) -> str:
    import re
    if not url:
        return ""
    # Strip tracking params from LinkedIn / ATS URLs
    if "linkedin.com/jobs/" in url:
        m = re.search(r'(\d{8,})', url)
        if m:
            return f"https://www.linkedin.com/jobs/view/{m.group(1)}/"
    # General URL clean: strip tracking query params
    url_clean = url.split("?")[0].split("#")[0].rstrip("/")
    return url_clean


class TrackerSync:
    """Syncs discovered/evaluated jobs to JSON queue and CSV export for Excel tracker."""

    def __init__(self, jobs_path: Optional[Path] = None):
        ensure_data_dirs()
        self.jobs_path = Path(jobs_path or JOBS_PATH)

    def load(self) -> List[JobListing]:
        return load_jobs(str(self.jobs_path))

    def save(self, jobs: List[JobListing]) -> None:
        save_jobs(str(self.jobs_path), jobs)

    def upsert(self, new_jobs: List[JobListing]) -> List[JobListing]:
        loaded = self.load()
        existing = {}

        for j in loaded:
            norm_url = normalize_job_url(j.job_url) or j.job_url
            comp_role = (j.company.strip().lower(), j.role_title.strip().lower())
            existing[norm_url] = j
            if comp_role != ("", ""):
                existing[comp_role] = j

        for job in new_jobs:
            norm_url = normalize_job_url(job.job_url) or job.job_url
            comp_role = (job.company.strip().lower(), job.role_title.strip().lower())

            prev = existing.get(norm_url) or existing.get(comp_role)
            if prev:
                job.job_id = prev.job_id or job.job_id
                job.job_url = prev.job_url or job.job_url
                # Unconditionally preserve applied / closed / non-new status
                if prev.status in (ApplicationStatus.APPLIED, ApplicationStatus.CLOSED, ApplicationStatus.INTERVIEW, ApplicationStatus.OFFER, ApplicationStatus.REJECTED):
                    job.status = prev.status
                    job.notes = prev.notes
                    job.attempt_count = prev.attempt_count
                elif job.status == ApplicationStatus.NEW:
                    job.status = prev.status

            existing[norm_url] = job
            if comp_role != ("", ""):
                existing[comp_role] = job

        # Return unique JobListing objects
        unique_jobs = list({id(j): j for j in existing.values()}.values())
        self.save(unique_jobs)
        return unique_jobs

    def export_csv(self, output_path: Optional[Path] = None) -> Path:
        output_path = output_path or self.jobs_path.parent / "jobs_export.csv"
        jobs = self.load()
        headers = [
            "Job ID", "Region", "Company", "Role Title", "Role Family", "Location", "Work Arrangement",
            "Portal", "Job URL", "Date Posted", "Age (Days)", "Priority Company?", "Routing",
            "Match Score", "Status", "Application Date", "Current CTC (LPA)", "Expected CTC (LPA)",
            "Resume Version", "Referral Contact", "Referral Status", "Follow-up Date", "Notes",
        ]

        evaluator = EvaluationAgent()
        today = date.today()
        rows = []
        for idx, job in enumerate(jobs, start=1):
            ev = evaluator.evaluate(job, today=today)
            age = ev.age_days if ev.age_days is not None else ""
            rows.append([
                job.job_id or f"JOB-{idx:03d}",
                getattr(job, "region", "India"),
                job.company,
                job.role_title,
                job.role_family,
                job.location,
                job.work_arrangement,
                job.portal,
                job.job_url,
                job.date_posted.isoformat() if job.date_posted else "",
                age,
                "Yes" if ev.is_priority_company else "No",
                ev.routing.value,
                f"{ev.match_score:.0%}",
                job.status.value,
                "",
                ev.current_ctc,
                ev.expected_ctc,
                "Pushp_Raj_Resume_Revised.pdf",
                "",
                "Not needed",
                "",
                job.notes,
            ])

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

        self.export_excel(headers, rows)
        self.sync_live_google_sheet(headers, rows)
        return output_path

    def sync_live_google_sheet(self, headers: List[str], rows: List[List[Any]]) -> bool:
        import os
        sheet_url_or_id = os.environ.get("GOOGLE_SHEET_URL") or os.environ.get("GOOGLE_SHEET_ID")
        if not sheet_url_or_id:
            env_path = Path(__file__).resolve().parent.parent / ".env"
            if env_path.exists():
                for line in env_path.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if line.startswith("GOOGLE_SHEET_URL=") or line.startswith("GOOGLE_SHEET_ID="):
                        sheet_url_or_id = line.split("=", 1)[1].strip().strip('"\'')
                        break

        if not sheet_url_or_id:
            return False

        try:
            import requests
            if "script.google.com" in sheet_url_or_id:
                response = requests.post(sheet_url_or_id, json={"headers": headers, "rows": rows}, allow_redirects=True, timeout=20)
                if response.status_code == 200 or "SUCCESS" in response.text:
                    print(f"📊 [Live Google Sheet] Successfully updated online sheet via Web App!")
                    return True

            import gspread
            gc = None
            service_account_path = Path(__file__).resolve().parent.parent / "google_service_account.json"
            if service_account_path.exists():
                gc = gspread.service_account(filename=str(service_account_path))
            else:
                try:
                    gc = gspread.oauth()
                except Exception:
                    pass

            if not gc:
                return False

            if "docs.google.com" in sheet_url_or_id:
                sh = gc.open_by_url(sheet_url_or_id)
            else:
                sh = gc.open_by_key(sheet_url_or_id)

            worksheet = sh.sheet1
            worksheet.clear()
            worksheet.update([headers] + rows)
            print(f"📊 [Live Google Sheet] Successfully updated online sheet: {sh.url}")
            return True
        except Exception as exc:
            print(f"⚠️ Live Google Sheet sync notice: {exc}")
            return False

    def export_excel(self, headers: List[str], rows: List[List[Any]]) -> Path:
        xlsx_path = Path(TRACKER_PATH)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Job Applications"
            ws.views.sheetView[0].showGridLines = True

            ws.append(headers)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row in rows:
                ws.append(row)

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

            wb.save(xlsx_path)
            return xlsx_path
        except Exception:
            return xlsx_path

    def summary(self) -> dict:
        jobs = self.load()
        evaluator = EvaluationAgent()
        results = evaluator.evaluate_batch(jobs)

        region_counts = {}
        for j in jobs:
            reg = getattr(j, "region", "India")
            region_counts[reg] = region_counts.get(reg, 0) + 1

        return {
            "total": len(jobs),
            "regions": region_counts,
            "auto_apply": sum(1 for r in results if r.routing == RoutingDecision.AUTO_APPLY),
            "hold_for_review": sum(1 for r in results if r.routing == RoutingDecision.HOLD_FOR_REVIEW),
            "skip": sum(1 for r in results if r.routing == RoutingDecision.SKIP),
            "applied": sum(1 for j in jobs if j.status == ApplicationStatus.APPLIED),
        }

    def apply_evaluations(self, evaluations: List[EvaluationResult]) -> None:
        jobs = self.load()
        by_url = {j.job_url: j for j in jobs}
        for ev in evaluations:
            job = by_url.get(ev.job.job_url)
            if not job:
                continue
            job.match_score = ev.match_score
            job.routing = ev.routing
            # NEVER overwrite terminal / active statuses (Applied, Closed, Interview, Offer, Rejected)
            if job.status not in (ApplicationStatus.APPLIED, ApplicationStatus.CLOSED, ApplicationStatus.INTERVIEW, ApplicationStatus.OFFER, ApplicationStatus.REJECTED):
                if ev.routing == RoutingDecision.HOLD_FOR_REVIEW:
                    job.status = ApplicationStatus.HELD_FOR_REVIEW
                elif ev.routing == RoutingDecision.AUTO_APPLY:
                    job.status = ApplicationStatus.READY_TO_APPLY
        self.save(list(by_url.values()))
